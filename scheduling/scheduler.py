from datetime import date, timedelta
import calendar
import random
from collections import defaultdict
from .models import Employee, Location, Holiday, EmployeeOffDay, Schedule
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


class ScheduleGenerator:
    SHIFTS = ['10AM-7PM', '1PM-10PM', '3PM-12AM']
    PREFERRED_FEMALE_SHIFTS = ['10AM-7PM', '1PM-10PM']

    def __init__(self, month, year, locations):
        self.month = month
        self.year = year
        self.locations = list(locations)
        self.employees = list(Employee.objects.filter(is_active=True))
        self.holidays = set(Holiday.objects.filter(
            date__month=month, date__year=year
        ).values_list('date', flat=True))
        self.off_days = defaultdict(set)

        # Load employee off days
        for off_day in EmployeeOffDay.objects.filter(
                date__month=month, date__year=year
        ).select_related('employee'):
            self.off_days[off_day.employee].add(off_day.date)

    def generate(self):
        """Generate the complete monthly schedule"""
        if not self.employees:
            raise ValueError("No active employees found")
        if not self.locations:
            raise ValueError("No locations selected")

        schedules = []
        month_days = calendar.monthrange(self.year, self.month)[1]

        # Track employee workload for balancing
        employee_shifts = defaultdict(int)

        for day in range(1, month_days + 1):
            current_date = date(self.year, self.month, day)

            # Skip holidays
            if current_date in self.holidays:
                continue

            # Weekends have no off days (as per requirement)
            is_weekend = current_date.weekday() >= 5

            # Track employees assigned today to prevent double-booking
            assigned_today = set()

            for location in self.locations:
                for shift in self.SHIFTS:
                    employee = self._assign_employee_to_shift(
                        current_date, shift, location, employee_shifts, is_weekend, assigned_today
                    )

                    if employee:
                        schedule = Schedule(
                            employee=employee,
                            location=location,
                            date=current_date,
                            shift=shift
                        )
                        schedules.append(schedule)
                        employee_shifts[employee] += 1
                        assigned_today.add(employee)
                    else:
                        # Log when we can't assign anyone to a shift
                        print(
                            f"Warning: Could not assign employee to {location.name} on {current_date} for {shift} shift")

        # Bulk create all schedules
        if schedules:
            Schedule.objects.bulk_create(schedules)
        return len(schedules)

    def _assign_employee_to_shift(self, current_date, shift, location, employee_shifts, is_weekend, assigned_today):
        """Assign the best available employee to a shift"""
        available_employees = []

        for employee in self.employees:
            # Skip if employee has off day (unless weekend)
            if not is_weekend and current_date in self.off_days[employee]:
                continue

            # Skip if employee already assigned today
            if employee in assigned_today:
                continue

            # Skip if employee already scheduled for this date
            if Schedule.objects.filter(employee=employee, date=current_date).exists():
                continue

            # Gender preference for shifts
            if employee.gender == 'F' and shift not in self.PREFERRED_FEMALE_SHIFTS:
                # Female employees can work 3PM-12AM shift but it's not preferred
                # Add them with lower priority
                available_employees.append((employee, employee_shifts[employee], 1))
            else:
                # Preferred assignment
                available_employees.append((employee, employee_shifts[employee], 0))

        if not available_employees:
            return None

        # Sort by: preference (0 is better), then by workload (fewer shifts is better), then random
        available_employees.sort(key=lambda x: (x[2], x[1], random.random()))

        return available_employees[0][0]

    def export_to_excel(self):
        """Export schedule to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Schedule {calendar.month_name[self.month]} {self.year}"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ['Date', 'Day', 'Location', 'Shift', 'Employee', 'Gender']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Data
        schedules = Schedule.objects.filter(
            date__month=self.month,
            date__year=self.year
        ).select_related('employee', 'location').order_by('date', 'location', 'shift')

        row = 2
        for schedule in schedules:
            day_name = schedule.date.strftime('%A')

            data = [
                schedule.date.strftime('%Y-%m-%d'),
                day_name,
                str(schedule.location),
                schedule.shift,
                schedule.employee.name,
                'Female' if schedule.employee.gender == 'F' else 'Male'
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border

                # Weekend highlighting
                if schedule.date.weekday() >= 5:
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        self._create_summary_sheet(summary_ws)

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file.getvalue()

    def _create_summary_sheet(self, ws):
        """Create summary statistics sheet"""
        ws.title = "Summary"

        # Headers
        ws['A1'] = 'Employee Statistics'
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = 'Employee'
        ws['B3'] = 'Total Shifts'
        ws['C3'] = 'Weekend Shifts'
        ws['D3'] = 'Weekday Shifts'

        # Style headers
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws[cell].font = Font(bold=True)

        # Employee statistics
        row = 4
        schedules = Schedule.objects.filter(
            date__month=self.month,
            date__year=self.year
        ).select_related('employee')

        employee_stats = defaultdict(lambda: {'total': 0, 'weekend': 0, 'weekday': 0})

        for schedule in schedules:
            emp = schedule.employee.name
            employee_stats[emp]['total'] += 1
            if schedule.date.weekday() >= 5:
                employee_stats[emp]['weekend'] += 1
            else:
                employee_stats[emp]['weekday'] += 1

        for emp_name, stats in sorted(employee_stats.items()):
            ws[f'A{row}'] = emp_name
            ws[f'B{row}'] = stats['total']
            ws[f'C{row}'] = stats['weekend']
            ws[f'D{row}'] = stats['weekday']
            row += 1
